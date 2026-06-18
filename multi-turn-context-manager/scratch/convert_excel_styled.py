import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import unicodedata
import os

def get_visual_width(text):
    width = 0
    for char in str(text):
        if unicodedata.east_asian_width(char) in ('W', 'F', 'A'):
            width += 2
        else:
            width += 1
    return width

csv_path = r'D:\VJ\Tro-li-ao-Javis\multi-turn-context-manager\reports\tests\test_summary_06_18.csv'
xlsx_path = r'D:\VJ\Tro-li-ao-Javis\multi-turn-context-manager\reports\tests\test_summary_06_18.xlsx'

print(f"Reading CSV from: {csv_path}")
df = pd.read_csv(csv_path, encoding='utf-8')

print(f"Generating styled Excel workbook...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Summary"

# Set grid lines visible explicitly
ws.views.sheetView[0].showGridLines = True

# Typography and Styling definitions
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid') # Slate dark blue
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')

pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') # Soft green
pass_font = Font(name='Segoe UI', size=11, bold=True, color='155724')

fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # Soft red
fail_font = Font(name='Segoe UI', size=11, bold=True, color='721C24')

normal_font = Font(name='Segoe UI', size=10, bold=False, color='000000')
border_side = Side(border_style='thin', color='BDC3C7')
cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

# Write header row
headers = list(df.columns)
ws.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = cell_border
ws.row_dimensions[1].height = 30

# Write data rows
import re
for row_idx, row in enumerate(df.values, 2):
    # Convert all values to strings and format None as empty string
    row_list = ["" if pd.isna(val) else str(val) for val in row]
    ws.append(row_list)
    for col_idx, val in enumerate(row_list, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = normal_font
        cell.border = cell_border
        
        # Alignment mapping
        col_name = headers[col_idx-1]
        
        # Format line breaks for text columns
        if col_name in [
            'Full Dialogue Flow (Nguyên bản tiếng Nhật)', 
            'Full Dialogue Flow (Bản dịch tiếng Việt)', 
            'Expected Result (Ground Truth)', 
            'Actual Result (Captured Answer)'
        ] and val:
            parts = val.split(" | ")
            formatted_parts = []
            for p in parts:
                p_formatted = re.sub(r'\s{2,}(?=(?:[^\s:：]+[:：]|\d+\.|\-\s))', '\n', p)
                formatted_parts.append(p_formatted)
            val = "\n\n".join(formatted_parts)
            cell.value = val
            
        if col_name in ['Version', 'Category', 'Scenario ID', 'Total Turns', 'Status']:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
        # Status highlights
        if col_name == 'Status':
            if val == 'PASS':
                cell.fill = pass_fill
                cell.font = pass_font
            elif val == 'FAIL':
                cell.fill = fail_fill
                cell.font = fail_font

# Dynamic Column Width and wrapping adjustments
for i, col in enumerate(df.columns):
    max_width = get_visual_width(col)
    for val in df[col]:
        if not pd.isna(val):
            max_width = max(max_width, get_visual_width(val))
    
    adjusted_width = max(12, min(max_width + 4, 100))
    # Standard reading widths for long paragraphs to prevent wide columns
    if col in [
        'Full Dialogue Flow (Nguyên bản tiếng Nhật)', 
        'Full Dialogue Flow (Bản dịch tiếng Việt)', 
        'Expected Result (Ground Truth)', 
        'Actual Result (Captured Answer)', 
        'Technical Validation (Tiếng Việt)', 
        'Functional Evaluation (Tiếng Việt)'
    ]:
        adjusted_width = 45
        
    column_letter = chr(65 + i) if i < 26 else f"{chr(64 + i // 26)}{chr(65 + i % 26)}"
    ws.column_dimensions[column_letter].width = adjusted_width

print(f"Saving styled Excel to: {xlsx_path}")
wb.save(xlsx_path)
print("Conversion completed successfully.")
