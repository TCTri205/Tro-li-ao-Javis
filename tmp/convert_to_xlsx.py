import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define file paths
md_file_path = r"D:\VJ\Tro-li-ao-Javis\tmp\gemini_result.md"
xlsx_file_path = r"D:\VJ\Tro-li-ao-Javis\tmp\gemini_result.xlsx"

# Read markdown content
with open(md_file_path, "r", encoding="utf-8") as f:
    content = f.read()

# Regex pattern to identify row start
pattern = r"Row\s+(\d+)Case\s+test:"
matches = list(re.finditer(pattern, content))

data = []

# Headers to locate inside each block
h1 = "Case test:"
h2 = "Transcript gốc (Data test):"
h3 = "Transcript mới (New Transcript):"
h4 = "Kì vọng cho Transcript mới:"

# Parse each match block
for i, match in enumerate(matches):
    start_pos = match.start()
    end_pos = matches[i+1].start() if i + 1 < len(matches) else len(content)
    row_num = int(match.group(1))
    
    block = content[start_pos:end_pos]
    
    # Locate header positions within the block
    pos_h1 = block.find(h1)
    pos_h2 = block.find(h2)
    pos_h3 = block.find(h3)
    pos_h4 = block.find(h4)
    
    # Check if all headers are found
    if pos_h1 != -1 and pos_h2 != -1 and pos_h3 != -1 and pos_h4 != -1:
        case_test = block[pos_h1 + len(h1) : pos_h2].strip()
        transcript_goc = block[pos_h2 + len(h2) : pos_h3].strip()
        transcript_moi = block[pos_h3 + len(h3) : pos_h4].strip()
        ki_vong = block[pos_h4 + len(h4) :].strip()
    else:
        # Fallback if parsing fails for some reason
        case_test = "Parsing error"
        transcript_goc = block
        transcript_moi = ""
        ki_vong = ""
        
    data.append({
        "Row Number": row_num,
        "Case Test": case_test,
        "Transcript gốc (Data test)": transcript_goc,
        "Transcript mới (New Transcript)": transcript_moi,
        "Kì vọng cho Transcript mới": ki_vong
    })

# Convert to DataFrame
df = pd.DataFrame(data)

# Save to Excel
df.to_excel(xlsx_file_path, index=False)

# Format Excel using openpyxl for a professional look
wb = openpyxl.load_workbook(xlsx_file_path)
ws = wb.active
ws.title = "Test Cases"

# Colors and Fonts
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Segoe UI", size=10)
bold_font = Font(name="Segoe UI", size=10, bold=True)

# Borders
thin_side = Side(border_style="thin", color="D9D9D9")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Alignment
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="top")

# Set formatting on header row (row 1)
ws.row_dimensions[1].height = 28
for col_num in range(1, 6):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_all

# Set formatting on data rows
for r_idx in range(2, len(data) + 2):
    # Calculate a reasonable row height based on the number of newlines to avoid clipping
    max_lines = 1
    for col_idx in range(1, 6):
        cell_val = ws.cell(row=r_idx, column=col_idx).value
        if cell_val:
            lines = str(cell_val).count("\n") + 1
            max_lines = max(max_lines, lines)
    ws.row_dimensions[r_idx].height = 18 + (max_lines - 1) * 14

    for col_idx in range(1, 6):
        cell = ws.cell(row=r_idx, column=col_idx)
        cell.font = cell_font
        cell.border = border_all
        
        if col_idx == 1:
            cell.alignment = center_alignment
            cell.font = bold_font
        else:
            cell.alignment = cell_alignment

# Adjust Column Widths
column_widths = {
    1: 12,  # Row Number
    2: 30,  # Case Test
    3: 50,  # Transcript gốc
    4: 50,  # Transcript mới
    5: 50   # Kì vọng
}

for col_idx, width in column_widths.items():
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

# Enable grid lines visibility
ws.views.sheetView[0].showGridLines = True

# Save formatted workbook
wb.save(xlsx_file_path)
print(f"Successfully converted and styled Excel file at: {xlsx_file_path}")
